from openpyxl import load_workbook


def showAllSheets(file_path):
    book = load_workbook(file_path)
    print(book.sheetnames)
def showMatrix(matrix):
    for row in matrix:
        print(row)

file_path = '../Excel/Immoexcel.xlsx'
book = load_workbook(file_path)
# sheet = book.active
sheet = book['Company']

new_data = [10,"TRUC",30,40,50]
matrix = []
for row in sheet['d6':'f6']:
    matrix.append(row)
    for index, cell in enumerate(row):
        cell.value = new_data[index]
book.save('newReport.xlsx')
